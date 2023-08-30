import tkinter as tk
from tkinter import filedialog, messagebox, Menu
import openpyxl
import logging
import requests
import subprocess
import os
from tkinter import ttk
import threading
import json
import socket

#developed by Sindre under the MIT license


# check if there is internet connection 
def check_internett_connection():
    try:
        socket.create_connection(("www.google.com", 80))
        return True
    except OSError:
        pass
    return False

latest_version = None
# set the version and the version URL and the download URL
CURRENT_VERSION = "1.1.0" 
'''
VERSION_URL = "https://raw.githubusercontent.com/BeeTwenty/barcode_scanner/master/version.txt"

if check_internett_connection():
    
        response = requests.get(VERSION_URL)
        latest_version = response.text.strip()
else: 
    logging.info("No internet connection. Skipping update check.")

# set the download URL, preferences file and debug mode
DOWNLOAD_URL = "https://github.com/BeeTwenty/barcode_scanner/releases/download/{}/BarcodeSetup.exe".format(latest_version)
'''
PREFERENCES_FILE = "preferences.json"
DEBUG_MODE = False
debug_mode = None
DEFAULT_DEBUG_MODE = False




'''
def download_and_install_update():
    # Create a new window for the progress bar
    progress_window = tk.Toplevel() # Create a new window
    progress_window.title("Update Progress") # Set the title
    progress_window.resizable(False, False)# Disable resizing
    progress_window.geometry("400x150") # Set the size
    progress_window.grab_set()# Make the window modal
    progress_window.focus_set() # Make the window modal 
    
    label = tk.Label(progress_window, text="Downloading update...") # Create a label and add it to the window 
    label.pack(padx=10, pady=10) # Set the padding

    label_file = tk.Label(progress_window, text="Downloading file...")
    label_file.pack(padx=10, pady=5)

    progress_bar = ttk.Progressbar(progress_window, length=300, mode="determinate") # Create a progress bar and add it to the window 
    progress_bar.pack(padx=10, pady=10)

# Create a thread for downloading the update and installing it 
    def download_thread():
        temp_file_path = "BarcodeSetup.exe" # Set the path for the temporary file
        logging.info("Downloading update...") # Log the download
        try: # Try to download the update
            response = requests.get(DOWNLOAD_URL, stream=True) # Download the update and stream it to the temporary file path 
            if response.status_code == 200: # Check if the download was successful 
                with open(temp_file_path, 'wb') as f: # Open the temporary file path and write the update to it 
                    total_size = int(response.headers.get('content-length', 0)) # Get the total size of the update 
                    block_size = 1024 # Set the block size to 1024 bytes 
                    progress = 0 # Set the progress to 0


                    for chunk in response.iter_content(chunk_size=1024): # Iterate over the update in chunks of 1024 bytes 
                        f.write(chunk) # Write the chunk to the temporary file path 
                        progress += block_size # Add the block size to the progress 

                        progress_bar["value"] = (progress / total_size) * 100 # Update the progress bar 
                        progress_window.update_idletasks() # Update the window 
                        label_file["text"] = "Downloading file: {:.1f}%".format((progress / total_size) * 100) # Update the label 


                f.close() # Close the file 
                logging.info("Download completed.")
                if os.path.isfile(temp_file_path): # Check if the temporary file exists 
                    progress_window.destroy() # Destroy the progress window 
                    # Run the installer and close the program
                    logging.info("Installing update...")   
                    subprocess.check_call([temp_file_path]) # Run the installer 
                    os.remove(temp_file_path) # Remove the temporary file 
                    

                    if messagebox.askyesno("Update", "Update installed successfully. Do you want to exit and restart the application?"):
                        window.quit() # Quit the program if the user clicks yes 
                        logging.info("Update installed successfully. Restarting application.")
                    logging.info("Update installed successfully")
                else:
                    messagebox.showerror("Update Error", "Update installation failed.")
                    logging.error("Update installation failed.")
            else:
                messagebox.showerror("Update Error", "Failed to download update. Reason: {}".format(response.reason))
                logging.error("Failed to download update. Reason: {}".format(response.reason))

        except requests.exceptions.RequestException as e:
            messagebox.showerror("Update Error", "Error occurred while downloading update: {} ".format(response.reason))
            logging.error("Error occurred while downloading update: {} ".format(response.reason))


        except subprocess.CalledProcessError as e:
            messagebox.showerror("Update Error", "Error occurred while installing update: {} ".format(response.reason))
            logging.error("Error occurred while installing update: {} ".format(response.reason))

        except Exception as e:
            messagebox.showerror("Update Error", "An error occurred: {} ".format(response.reason))
            logging.error("An error occurred: {} ".format(response.reason))
    # Start the download thread
    thread = threading.Thread(target=download_thread) # Create a thread for downloading the update 
    thread.start() # Start the thread 
    
    


def check_updates():

    try:
        if check_internett_connection():

            logging.info("Checking for updates...")
            try:
                # Fetch the latest version from the version URL
                response = requests.get(VERSION_URL)
                latest_version = response.text.strip()
                logging.info(f"Latest version: {latest_version}")      
        

                # Compare the current version with the latest version
                if latest_version > CURRENT_VERSION:
                    d_response = messagebox.askquestion("Update Available", "A new version ({}) is available. Do you Want to download now?.".format(latest_version))
                    if d_response == "yes":
                        download_and_install_update() 
                    logging.info("Update available. Please update. ( {} )".format(latest_version))
        
                else:
                    messagebox.showinfo("Up to Date", "You have the latest version of the program.")
                    logging.info("Up to date. ( {} )".format(latest_version))

            except requests.exceptions.RequestException:
                messagebox.showerror("Error", "Failed to check for updates.")
                logging.error("Failed to check for updates.")
        else:
            
            logging.error("No Internet")
            messagebox.showerror("Error", "No Internet")
    except requests.exceptions.RequestException as e:
        messagebox.showerror("Error", "Failed to check for updates. Reaseon: {str(e)}")
        logging.error("Failed to check for updates.") 

def check_updates_at_start():
    load_preferences()
    setup_logging()
    logging.info("Barcode Scanner started. Version: {}".format(CURRENT_VERSION))
    if check_internett_connection():
        logging.info("Checking for updates...")
    
    try:
        if check_internett_connection():
            logging.info("Checking for updates...")
        
            # Fetch the latest version from the version URL
            response = requests.get(VERSION_URL)
            latest_version = response.text.strip()
            logging.info(f"Latest version: {latest_version}")      

            # Compare the current version with the latest version
            if latest_version > CURRENT_VERSION:
                d_response = messagebox.askquestion("Update Available", "A new version ({}) is available. Do you want to download now?".format(latest_version))
                if d_response == "yes":
                    download_and_install_update()
                logging.info("Update available. Please update. ({})".format(latest_version))
            else:
                logging.info("Up to date. ({})".format(latest_version))
        else:
            logging.info("No internet connection. Skipping update check.")
    except requests.exceptions.RequestException:
        messagebox.showerror("Error", "Failed to check for updates.")
        logging.error("Failed to check for updates.")
   
'''
# Mark barcode in Excel sheet
def mark_barcode_in_excel(barcode, workbook_path, barcode_column, box_value=None):
    try:
        workbook = openpyxl.load_workbook(workbook_path)
        sheet = workbook.active
        logging.info(f"Workbook opened: {workbook_path}")
        barcode_found = False

        # Loop through all cells in the barcode and box columns
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=7):
            tape, box = row[1].value, row[4].value  # Assuming 'TAPE' is in column B and 'C.BARC' is in column E

            if tape == barcode:
                if box_value:
                    if box == box_value:
                        row[1].fill = openpyxl.styles.PatternFill(start_color="00FF00", fill_type="solid")
                        barcode_found = True
                        logging.info(f"Barcode found in specified box: {barcode}")
                else:
                    row[1].fill = openpyxl.styles.PatternFill(start_color="00FF00", fill_type="solid")
                    barcode_found = True
                    logging.info(f"Barcode found: {barcode}")

        # Save the workbook
        workbook.save(workbook_path)
        workbook.close()
        logging.info(f"Workbook saved: {workbook_path} and closed.")

        if not barcode_found:
            messagebox.showerror("Error", "Barcode not found.")
            logging.error(f"Barcode not found: {barcode}")

    except FileNotFoundError:
        messagebox.showerror("Error", "Workbook not found.")
        logging.error(f"Workbook not found: {workbook_path}")

    except Exception as e:
        messagebox.showerror("Error", str(e))
        logging.error(f"Error marking barcode: {barcode}. Error: {str(e)}")

# create save_preferences function for debug mode
def save_preferences():
    global DEBUG_MODE
    DEBUG_MODE = debug_mode.get()
    logging.info(f"Debug mode: {DEBUG_MODE}")
    messagebox.showinfo("Debug Mode", f"Debug mode: {DEBUG_MODE}")

    # Save preferences to JSON file
    preferences = {"debug_mode": DEBUG_MODE}
    with open(PREFERENCES_FILE, "w") as file:
        json.dump(preferences, file)
    logging.info("Preferences saved.")

def load_preferences():
    global DEBUG_MODE
    try:
        with open(PREFERENCES_FILE, "r") as file:
            preferences = json.load(file)
            DEBUG_MODE = preferences.get("debug_mode", DEFAULT_DEBUG_MODE)
            logging.info(f"Loaded debug mode: {DEBUG_MODE}")
    except FileNotFoundError:
        logging.info("Preferences file not found. Using default settings.")
        DEBUG_MODE = DEFAULT_DEBUG_MODE
    except json.JSONDecodeError:
        logging.error("Error parsing preferences file. Using default settings.")
        DEBUG_MODE = DEFAULT_DEBUG_MODE

def setup_logging():
    log_file = "barcode_log.txt"
    logger = logging.getLogger()  # Get the root logger

    # Remove any existing handlers to avoid duplication
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)

    if DEBUG_MODE:
        logger.setLevel(logging.INFO)
    else:
        logger.setLevel(logging.ERROR)

    # Create a file handler and set its level to INFO
    file_handler = logging.FileHandler(log_file)
    file_handler.setLevel(logging.INFO)

    # Create a console handler and set its level to ERROR
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.ERROR)

    # Create a formatter and add it to the handlers
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)

    # Add the handlers to the logger
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

# Scan barcode from entry field and mark it in Excel sheet
def scan_barcode(event):
    barcode = barcode_entry.get()
    wb_path = workbook_entry.get()
    bc_column = column_entry.get()
    box_value = box_entry.get() if box_mode.get() else None
    mark_barcode_in_excel(barcode, wb_path, bc_column, box_value)
    barcode_entry.delete(0, tk.END)  # Clear the barcode entry field after scanning
    logging.info(f"Barcode scanned: {barcode}")

# Browse for workbook file
def browse_workbook():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
    workbook_entry.delete(0, tk.END)
    workbook_entry.insert(tk.END, file_path)
    logging.info(f"Workbook selected: {file_path}")

# Show about window with information about the program
def show_about_window():
    about_text = "Barcode Scanner\n\nVersion: {}\n\nDeveloped by: Sindre\n\nDescription: Enter a barcode to mark it as green in the Excel sheet.\n \n Note: Due to Windows Locking the Excel file when it is open, the program can't run with the file open.".format(CURRENT_VERSION)

    messagebox.showinfo("About", about_text) 
    logging.info("About window opened.")

def show_preference_window():
    # Create the preference window with option to activate/deactivate debug mode
    preference_window = tk.Toplevel(window)
    preference_window.title("Preferences")
    preference_window.geometry("300x100")
    preference_window.resizable(False, False)
    preference_window.iconbitmap("barcode.ico")
    preference_window.grab_set()  # Make the preference window the active window

    global debug_mode  # Declare debug_mode as a global variable

    # Create the debug mode checkbox
    debug_mode = tk.IntVar(value=DEBUG_MODE)
    debug_mode_checkbox = tk.Checkbutton(preference_window, text="Debug Mode", variable=debug_mode)
    debug_mode_checkbox.grid(row=0, column=0, padx=10, pady=10)

    # Load preferences
    load_preferences()

    # Create the save button
    save_button = tk.Button(preference_window, text="Save", command=save_preferences)
    save_button.grid(row=1, column=0, padx=10, pady=10)

    logging.info("Preferences window opened.")

# Create the main window
window = tk.Tk() # Create the main window
window.title("Barcode Scanner") # Set the window title 
window.geometry("400x350") # Set the window size 
tk.Tk.iconbitmap(window, default="barcode.ico") # Set the window icon
window.resizable(False, False) # Disable resizing of the window


logging.info("Main window created.")
# Create the menu bar
menu = Menu(window) # Create the menu bar
help = Menu(menu, tearoff=0) # Create the Help menu item
help.add_command(label="About", command=show_about_window)
#help.add_command(label="Update", command=check_updates) # Add About menu item to Help menu
menu.add_cascade(label="Help", menu=help) # Add Help menu to menu bar
options = Menu(menu, tearoff=0) # Create the Options menu item
options.add_command(label="Preferences", command=show_preference_window) # Add Preferences menu item to Options menu
menu.add_cascade(label="Options", menu=options) # Add Options menu to menu bar
window.config(menu=menu) # Add menu bar to window

# Create the GUI
label_workbook = tk.Label(window, text="Workbook Path:") # Create the workbook path label
label_workbook.pack(pady=10) # Add padding to the label to make it look better

# Create the workbook path entry field
workbook_entry = tk.Entry(window) # Create the workbook path entry field
workbook_entry.pack(padx=5)

# Create the browse button to browse for workbook file
browse_button = tk.Button(window, text="Browse", command=browse_workbook)
browse_button.pack(pady=5)

# Create the barcode column entry field 
label_column = tk.Label(window, text="Barcode Column:")
label_column.pack(pady=10)
column_entry = tk.Entry(window)
column_entry.pack()

# Add Box Mode Checkbox and Entry
box_mode = tk.IntVar()  # 0 for off, 1 for on
box_mode_checkbox = tk.Checkbutton(window, text="Box Mode", variable=box_mode)
box_mode_checkbox.pack(pady=10)

label_box = tk.Label(window, text="Enter Box:")
label_box.pack(pady=10)
box_entry = tk.Entry(window)
box_entry.pack()

# Create the barcode entry field 
label_barcode = tk.Label(window, text="Scan Barcode:")
label_barcode.pack(pady=10)
barcode_entry = tk.Entry(window)
barcode_entry.pack()



# Bind the Return key event to scan_barcode function
barcode_entry.bind("<Return>", scan_barcode)

#check_updates_at_start()
#check_updates_at_start()


# Run the main window
window.mainloop()